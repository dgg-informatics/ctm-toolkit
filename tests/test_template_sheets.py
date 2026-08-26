"""Keeps the v1 reference workbook in sync with the parser.

The intake template is now hand-authored (test-pt-data-v1.0.0.xlsx, with colors
and cell-validation the code cannot regenerate) rather than emitted by a script.
These tests assert the committed workbook still matches what the parser expects:
every finding sheet present, the canonical columns spelled exactly, the report
identity columns present, and the rows actually join.
"""
from pathlib import Path

import openpyxl

from ctm.transformers.normalize_manual import SHEET_NORMALIZERS

REFERENCE = Path(__file__).parent / "fixtures" / "test-pt-data-v1.0.0.xlsx"

CANONICAL = {
    "pt_uuid", "report_uuid", "biomarker", "variant_category", "protein_change",
    "cnv_call", "signature_level", "wildtype", "nucleotide_change",
}
REPORT_IDENTITY = {
    "report_uuid", "pt_uuid", "source", "test_name",
    "unique_test_id", "unique_test_id_source",
}


def test_reference_workbook_covers_every_sheet_the_parser_reads():
    wb = openpyxl.load_workbook(REFERENCE, read_only=True)
    missing = set(SHEET_NORMALIZERS) - set(wb.sheetnames)
    assert not missing, f"reference workbook is missing {sorted(missing)}"


def test_findings_sheets_spell_the_canonical_columns_correctly():
    """extra='allow' means a typo'd header passes silently into `raw` and leaves
    the real field empty, so assert the canonical block is present and exact."""
    wb = openpyxl.load_workbook(REFERENCE, read_only=True)
    for sheet in SHEET_NORMALIZERS:
        headers = {c.value for c in wb[sheet][1] if c.value is not None}
        assert CANONICAL <= headers, f"{sheet}: missing/misspelled {sorted(CANONICAL - headers)}"


def test_report_metadata_has_the_identity_columns():
    wb = openpyxl.load_workbook(REFERENCE, read_only=True)
    headers = {c.value for c in wb["report_metadata"][1] if c.value is not None}
    assert REPORT_IDENTITY <= headers, f"report_metadata missing {sorted(REPORT_IDENTITY - headers)}"


def test_reference_workbook_is_parseable_and_joined():
    """Rows must actually join — orphaned pt_uuids are dropped without warning."""
    from ctm.transformers.excel_reader import read_and_normalize

    patients, _metadata, findings = read_and_normalize(REFERENCE)
    assert patients, "no patients parsed"
    assert findings, "findings present in the workbook but none survived the join"
